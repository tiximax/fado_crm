# 📊 FADO CRM - Export/Import Service
# Xuất/nhập dữ liệu Excel & CSV chuyên nghiệp! 🚀

from typing import List, Dict, Any, Optional, BinaryIO
import pandas as pd
import io
from datetime import datetime, date
from sqlalchemy.orm import Session
from sqlalchemy import text
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from models import KhachHang, SanPham, DonHang, ChiTietDonHang, LichSuLienHe, TrangThaiDonHang, LoaiKhachHang
from logging_config import app_logger

class ExportImportService:
    def __init__(self):
        self.db_session = None
        app_logger.info("📊 Export/Import service initialized")

    def set_session(self, db: Session):
        """🔌 Set database session"""
        self.db_session = db

    # 📤 EXPORT FUNCTIONS
    def export_customers_to_excel(self, filters: Optional[Dict[str, Any]] = None) -> bytes:
        """👥 Export customers to Excel"""
        try:
            query = self.db_session.query(KhachHang)

            # Apply filters
            if filters:
                if filters.get("customer_type"):
                    query = query.filter(KhachHang.loai_khach == filters["customer_type"])
                if filters.get("created_from"):
                    query = query.filter(KhachHang.ngay_tao >= filters["created_from"])
                if filters.get("created_to"):
                    query = query.filter(KhachHang.ngay_tao <= filters["created_to"])

            customers = query.all()

            # Create DataFrame
            data = []
            for customer in customers:
                data.append({
                    "ID": customer.id,
                    "Họ tên": customer.ho_ten,
                    "Email": customer.email,
                    "Số điện thoại": customer.so_dien_thoai,
                    "Địa chỉ": customer.dia_chi,
                    "Loại khách hàng": customer.loai_khach.value,
                    "Tổng tiền đã mua": customer.tong_tien_da_mua,
                    "Số đơn thành công": customer.so_don_thanh_cong,
                    "Ngày tạo": customer.ngay_tao.strftime("%d/%m/%Y %H:%M") if customer.ngay_tao else "",
                    "Ngày cập nhật": customer.ngay_cap_nhat.strftime("%d/%m/%Y %H:%M") if customer.ngay_cap_nhat else ""
                })

            df = pd.DataFrame(data)

            # Create Excel with styling
            return self._create_styled_excel(df, "Danh sách khách hàng", "KHACH_HANG")

        except Exception as e:
            app_logger.error(f"❌ Error exporting customers to Excel: {str(e)}")
            raise

    def export_products_to_excel(self, filters: Optional[Dict[str, Any]] = None) -> bytes:
        """🛍️ Export products to Excel"""
        try:
            query = self.db_session.query(SanPham)

            # Apply filters
            if filters:
                if filters.get("category"):
                    query = query.filter(SanPham.danh_muc.ilike(f"%{filters['category']}%"))
                if filters.get("country"):
                    query = query.filter(SanPham.quoc_gia_nguon.ilike(f"%{filters['country']}%"))
                if filters.get("min_price"):
                    query = query.filter(SanPham.gia_ban >= filters["min_price"])
                if filters.get("max_price"):
                    query = query.filter(SanPham.gia_ban <= filters["max_price"])

            products = query.all()

            # Create DataFrame
            data = []
            for product in products:
                data.append({
                    "ID": product.id,
                    "Tên sản phẩm": product.ten_san_pham,
                    "Mô tả": product.mo_ta,
                    "Danh mục": product.danh_muc,
                    "Giá bán": product.gia_ban,
                    "Trọng lượng (kg)": product.trong_luong,
                    "Quốc gia nguồn": product.quoc_gia_nguon,
                    "URL hình ảnh": product.hinh_anh_url,
                    "Ngày tạo": product.ngay_tao.strftime("%d/%m/%Y %H:%M") if product.ngay_tao else "",
                    "Ngày cập nhật": product.ngay_cap_nhat.strftime("%d/%m/%Y %H:%M") if product.ngay_cap_nhat else ""
                })

            df = pd.DataFrame(data)

            return self._create_styled_excel(df, "Danh sách sản phẩm", "SAN_PHAM")

        except Exception as e:
            app_logger.error(f"❌ Error exporting products to Excel: {str(e)}")
            raise

    def export_orders_to_excel(self, filters: Optional[Dict[str, Any]] = None) -> bytes:
        """📋 Export orders to Excel"""
        try:
            query = self.db_session.query(DonHang).join(KhachHang, isouter=True)

            # Apply filters
            if filters:
                if filters.get("status"):
                    if isinstance(filters["status"], list):
                        query = query.filter(DonHang.trang_thai.in_(filters["status"]))
                    else:
                        query = query.filter(DonHang.trang_thai == filters["status"])
                if filters.get("created_from"):
                    query = query.filter(DonHang.ngay_tao >= filters["created_from"])
                if filters.get("created_to"):
                    query = query.filter(DonHang.ngay_tao <= filters["created_to"])

            orders = query.all()

            # Create DataFrame
            data = []
            for order in orders:
                data.append({
                    "ID": order.id,
                    "Mã đơn hàng": order.ma_don_hang,
                    "Tên khách hàng": order.khach_hang.ho_ten if order.khach_hang else "N/A",
                    "Email khách hàng": order.khach_hang.email if order.khach_hang else "N/A",
                    "Trạng thái": order.trang_thai.value,
                    "Tổng tiền": order.tong_tien,
                    "Phí ship": order.phi_ship,
                    "Ghi chú": order.ghi_chu,
                    "Mã vận đơn": order.ma_van_don,
                    "Ngày tạo": order.ngay_tao.strftime("%d/%m/%Y %H:%M") if order.ngay_tao else "",
                    "Ngày giao hàng": order.ngay_giao_hang.strftime("%d/%m/%Y") if order.ngay_giao_hang else ""
                })

            df = pd.DataFrame(data)

            return self._create_styled_excel(df, "Danh sách đơn hàng", "DON_HANG")

        except Exception as e:
            app_logger.error(f"❌ Error exporting orders to Excel: {str(e)}")
            raise

    def export_to_csv(self, entity_type: str, filters: Optional[Dict[str, Any]] = None) -> str:
        """📄 Export to CSV format"""
        try:
            if entity_type == "customers":
                query = self.db_session.query(KhachHang)

                if filters:
                    if filters.get("customer_type"):
                        query = query.filter(KhachHang.loai_khach == filters["customer_type"])

                customers = query.all()
                data = []
                for customer in customers:
                    data.append({
                        "ID": customer.id,
                        "Họ tên": customer.ho_ten,
                        "Email": customer.email,
                        "Số điện thoại": customer.so_dien_thoai,
                        "Địa chỉ": customer.dia_chi,
                        "Loại khách hàng": customer.loai_khach.value,
                        "Tổng tiền đã mua": customer.tong_tien_da_mua,
                        "Số đơn thành công": customer.so_don_thanh_cong,
                        "Ngày tạo": customer.ngay_tao.strftime("%d/%m/%Y %H:%M") if customer.ngay_tao else ""
                    })

            elif entity_type == "products":
                query = self.db_session.query(SanPham)

                if filters:
                    if filters.get("category"):
                        query = query.filter(SanPham.danh_muc.ilike(f"%{filters['category']}%"))

                products = query.all()
                data = []
                for product in products:
                    data.append({
                        "ID": product.id,
                        "Tên sản phẩm": product.ten_san_pham,
                        "Mô tả": product.mo_ta,
                        "Danh mục": product.danh_muc,
                        "Giá bán": product.gia_ban,
                        "Trọng lượng": product.trong_luong,
                        "Quốc gia nguồn": product.quoc_gia_nguon,
                        "Ngày tạo": product.ngay_tao.strftime("%d/%m/%Y %H:%M") if product.ngay_tao else ""
                    })

            elif entity_type == "orders":
                query = self.db_session.query(DonHang).join(KhachHang, isouter=True)

                if filters:
                    if filters.get("status"):
                        query = query.filter(DonHang.trang_thai == filters["status"])

                orders = query.all()
                data = []
                for order in orders:
                    data.append({
                        "ID": order.id,
                        "Mã đơn hàng": order.ma_don_hang,
                        "Tên khách hàng": order.khach_hang.ho_ten if order.khach_hang else "N/A",
                        "Trạng thái": order.trang_thai.value,
                        "Tổng tiền": order.tong_tien,
                        "Ngày tạo": order.ngay_tao.strftime("%d/%m/%Y %H:%M") if order.ngay_tao else ""
                    })

            df = pd.DataFrame(data)
            return df.to_csv(index=False, encoding='utf-8-sig')

        except Exception as e:
            app_logger.error(f"❌ Error exporting to CSV: {str(e)}")
            raise

    # 📥 IMPORT FUNCTIONS
    def import_customers_from_excel(self, file_content: bytes, user_id: int) -> Dict[str, Any]:
        """👥 Import customers from Excel"""
        try:
            # Read Excel file
            df = pd.read_excel(io.BytesIO(file_content))

            # Validate required columns
            required_columns = ["Họ tên", "Email", "Số điện thoại"]
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                return {
                    "success": False,
                    "error": f"Thiếu các cột bắt buộc: {', '.join(missing_columns)}"
                }

            success_count = 0
            error_count = 0
            errors = []

            for index, row in df.iterrows():
                try:
                    # Check if customer already exists
                    existing = self.db_session.query(KhachHang).filter_by(email=row["Email"]).first()

                    if existing:
                        errors.append(f"Dòng {index + 2}: Email {row['Email']} đã tồn tại")
                        error_count += 1
                        continue

                    # Create new customer
                    customer = KhachHang(
                        ho_ten=row["Họ tên"],
                        email=row["Email"],
                        so_dien_thoai=str(row["Số điện thoại"]) if pd.notna(row["Số điện thoại"]) else None,
                        dia_chi=row.get("Địa chỉ", ""),
                        loai_khach=LoaiKhachHang.TIEM_NANG,
                        nguoi_tao_id=user_id
                    )

                    self.db_session.add(customer)
                    success_count += 1

                except Exception as e:
                    errors.append(f"Dòng {index + 2}: {str(e)}")
                    error_count += 1

            self.db_session.commit()

            return {
                "success": True,
                "message": f"Import thành công {success_count} khách hàng, {error_count} lỗi",
                "success_count": success_count,
                "error_count": error_count,
                "errors": errors[:10]  # Limit to first 10 errors
            }

        except Exception as e:
            self.db_session.rollback()
            app_logger.error(f"❌ Error importing customers from Excel: {str(e)}")
            return {
                "success": False,
                "error": f"Lỗi import: {str(e)}"
            }

    def import_products_from_excel(self, file_content: bytes, user_id: int) -> Dict[str, Any]:
        """🛍️ Import products from Excel"""
        try:
            df = pd.read_excel(io.BytesIO(file_content))

            # Validate required columns
            required_columns = ["Tên sản phẩm", "Giá bán"]
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                return {
                    "success": False,
                    "error": f"Thiếu các cột bắt buộc: {', '.join(missing_columns)}"
                }

            success_count = 0
            error_count = 0
            errors = []

            for index, row in df.iterrows():
                try:
                    # Create new product
                    product = SanPham(
                        ten_san_pham=row["Tên sản phẩm"],
                        mo_ta=row.get("Mô tả", ""),
                        danh_muc=row.get("Danh mục", ""),
                        gia_ban=float(row["Giá bán"]),
                        trong_luong=float(row.get("Trọng lượng", 0)) if pd.notna(row.get("Trọng lượng")) else 0,
                        quoc_gia_nguon=row.get("Quốc gia nguồn", ""),
                        nguoi_tao_id=user_id
                    )

                    self.db_session.add(product)
                    success_count += 1

                except Exception as e:
                    errors.append(f"Dòng {index + 2}: {str(e)}")
                    error_count += 1

            self.db_session.commit()

            return {
                "success": True,
                "message": f"Import thành công {success_count} sản phẩm, {error_count} lỗi",
                "success_count": success_count,
                "error_count": error_count,
                "errors": errors[:10]
            }

        except Exception as e:
            self.db_session.rollback()
            app_logger.error(f"❌ Error importing products from Excel: {str(e)}")
            return {
                "success": False,
                "error": f"Lỗi import: {str(e)}"
            }

    # 🎨 HELPER FUNCTIONS
    def _create_styled_excel(self, df: pd.DataFrame, sheet_name: str, file_prefix: str) -> bytes:
        """🎨 Create styled Excel file"""
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Style header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col in worksheet[1]:
                col.font = header_font
                col.fill = header_fill
                col.alignment = header_alignment

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

        output.seek(0)
        return output.read()

    def get_export_stats(self) -> Dict[str, Any]:
        """📊 Get export statistics"""
        try:
            stats = {
                "total_customers": self.db_session.query(KhachHang).count(),
                "total_products": self.db_session.query(SanPham).count(),
                "total_orders": self.db_session.query(DonHang).count(),
                "export_formats": ["Excel (.xlsx)", "CSV (.csv)"],
                "import_formats": ["Excel (.xlsx)"],
                "last_export": datetime.now().strftime("%d/%m/%Y %H:%M")
            }

            return stats

        except Exception as e:
            app_logger.error(f"❌ Error getting export stats: {str(e)}")
            return {}

# 🌟 Global export service
export_service = ExportImportService()

print("📊 Export/Import service loaded successfully!")